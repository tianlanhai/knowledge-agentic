# -*- coding: utf-8 -*-
"""
上海宇羲伏天智能科技有限公司出品

文件级注释：Excel 文档加载器
内部逻辑：使用 openpyxl 解析 Excel 文件，提取单元格文本内容
说明：轻量级替代方案，支持 .xlsx 和 .xls 格式
"""

from typing import List
from langchain_core.documents import Document


class ExcelLoader:
    """
    类级注释：Excel 文档加载器
    使用 openpyxl 库解析 .xlsx 文件
    """

    def __init__(self, file_path: str):
        """
        函数级注释：初始化 Excel 加载器
        参数：
            file_path: Excel 文件路径
        """
        self.file_path = file_path

    def load(self) -> List[Document]:
        """
        函数级注释：加载 Excel 文件并提取文本内容
        返回值：Document 对象列表
        """
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl 包未安装。请运行: pip install openpyxl"
            )

        # 内部逻辑：打开 Excel 文件
        workbook = openpyxl.load_workbook(self.file_path, data_only=True)

        # 内部变量：存储所有工作表的文本内容
        full_text = []

        # 内部逻辑：遍历所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # 内部变量：当前工作表的文本
            sheet_text = []

            # 内部逻辑：添加工作表标题
            sheet_text.append(f"[工作表: {sheet_name}]")

            # 内部逻辑：遍历所有行
            for row in sheet.iter_rows(values_only=True):
                # 内部变量：过滤空值并转换为字符串
                row_values = [str(cell) if cell is not None else "" for cell in row]
                # 内部逻辑：跳过完全空的行
                if any(row_values):
                    sheet_text.append(" | ".join(row_values))

            # 内部逻辑：如果有内容，添加到总内容中
            if len(sheet_text) > 1:  # 除了标题外还有内容
                full_text.append("\n".join(sheet_text))

        # 内部逻辑：关闭工作簿
        workbook.close()

        # 内部逻辑：创建 Document 对象
        content = "\n\n".join(full_text)

        metadata = {
            "source": self.file_path,
            "file_path": self.file_path,
        }

        return [Document(page_content=content, metadata=metadata)]
